#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @date 2022/5/1
# @file My_test.py
import os
import openpyxl


def COUNT(path):
    digits = 0
    alpha = 0
    others = 0
    with open(path, 'r') as f:
        _ = f.readlines()
        print(_)
        for line in _:
            for word in line:
                if '0' <= word <= '9':
                    digits += 1
                elif 'a' <= word <= 'z' or 'A' <= word <= 'Z':
                    alpha += 1
                else:
                    others += 1
    print(digits, alpha, others)


def Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    dic = {}
    rowList = range(2, sheet.max_row)
    for row in rowList:
        dic[sheet.cell(row, 1).value] = sheet.cell(row, 3).value
    dicActor = {}
    for k, v in dic.items():
        for actor in v.split(','):
            if actor not in dicActor:
                dicActor[actor] = []
            dicActor[actor].append(k)

    for k, v in dicActor.items():
        print(k, '参演: ', ', '.join(v))
