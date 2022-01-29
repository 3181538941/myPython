#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @author LeoWang
# @date 2022/1/8
# @file excelTest.py
import openpyxl
import xlrd
import xlwt
import os


class Excel(object):
    """
    Excel类, 用来进行简单的Excel操作
    """

    def __init__(self, file):
        self.file = file  # 文件全称
        self.sheetName = None
        if not self.fileExist():
            self.createNewExcel()
            print("文件不存在, 已新建文件")
        self.wb = openpyxl.load_workbook(self.file)

    def save(self):
        self.wb.save(self.file)

    def fileExist(self):
        return os.path.isfile(self.file)
        # self.dirName = os.path.split('./' + self.file)[0]  # 文件父目录名
        # self.fileName = os.path.split(self.file)[1]  # 文件名称
        #
        # fileList = os.listdir(self.dirName)  # 文件所在目录下文件列表
        #
        # if self.fileName in fileList:
        #     return 1
        # else:
        #     return 0

    def createNewExcel(self, sheetName=None):
        """
        创建Excel文件方法
        :param sheetName: 工作表的名称, 默认为 Sheet1, 可传列表实现初始化多个工作表
        :return:
        """
        if sheetName is None:
            sheetName = 'Sheet1'
        self.sheetName = sheetName

        wb = openpyxl.Workbook()  # 实例化工作簿对象
        if isinstance(self.sheetName, str):
            sheet = wb.active  # 获取活动工作表
            sheet.title = self.sheetName  # 为工作表从新命名
        elif isinstance(self.sheetName, list):  # 依次为表重新命名
            sheet = wb.active
            sheet.title = self.sheetName[0]
            for eachSheet in self.sheetName[1:]:
                wb.create_sheet(eachSheet)
        else:
            raise TypeError

        wb.save(self.file)

    def getSheets(self):
        self.sheets = self.wb.sheetnames
        print(self.sheets)

    def openBySheet(self,sheetName):
        self.sheet = self.wb[sheetName]

    def write(self):
        pass

    def read(self):
        pass

    def rowAppend(self):
        pass

    def colAppend(self):
        pass


# file = Excel('excels/高中同学录取及分数.xls')
# print(file.wb.sheetnames)

# wb = xlrd.open_workbook('excels/高中同学录取及分数.xls')
# sheet = wb['Sheet1']
# print(sheet)
#
# print(sheet.name, sheet.nrows, sheet.ncols)
# print(sheet.col_values(1))

# file = Excel('excels/手机.xlsx')
# file.getSheets()
wb = openpyxl.load_workbook('excels/手机.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet2']
sheet.append([1])
# sheet = wb.worksheets[1]
# print(sheet.rows)
# for row in sheet.rows:
#     # 循环遍历每一个单元格
#     for cell in row[1:]:
#         # 获取单元格的内容
#         print(f"{str(cell.value):>5s}", end=',')
#     print()
# for row in range(1, 21):
#     for i in range(6):
#         col = chr(ord("A") + i)
#         # print(col,row)
#         sheet[f'{col}{row}'].value = f'{col}{row}'
wb.save('excels/手机.xlsx')
# shh = ['sh1', 'sh2', 'sh3']
# file.createNewExcel(sheetName=shh)


# os.system('del excels /Q')
# os.remove('excels/t2e2.xlsx')

# dir = os.path.split(r'.\test1.xlsx')[0]
# list = os.listdir(dir)
# print(list)
# print(dir)
# print(os.getcwd())
# wb=openpyxl.load_workbook('test1.xlsx')
# # wb = openpyxl.Workbook(r'F:\03_Important\Python\0a_project\clear_up_files\learn\excels\test2.xlsx')
# t = wb.sheetnames
# # wb.save(r'excels\test1.xlsx')
# print(t)

# sheet1 = wb['Sheet1']
# print(sheet1)
# sheet = wb.worksheets[0]
# print(sheet)

# with open(os.getcwd()+'\\test1.xlsx','r',encoding='utf-8') as xl:
#     # xl.write('test')
#     xl.seek(0)
#     st = xl.read()
#     print(str(st))


# from openpyxl import Workbook
# import datetime
# import time
# import locale
#
# wb = Workbook()  # 创建文件对象
#
# ws = wb.active  # 获取第一个sheet
#
# ws['A2'] = datetime.datetime.now()  # 写入一个当前时间
#
# # 写入一个自定义的时间格式
# locale.setlocale(locale.LC_CTYPE, 'chinese')
# ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())
#
# # Save the file
# wb.save("sample.xlsx")
